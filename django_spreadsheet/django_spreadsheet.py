import inspect
import mimetypes
from io import BytesIO

from django.http import HttpResponse
from django.views.generic import View

import openpyxl


class Worksheet(openpyxl.worksheet.worksheet.Worksheet):
    column_width_factor = 1.23

    def get_queryset(self):
        return self.model.objects.all()

    def fill(self):
        self._fill_headings()
        self._fill_body()
        self._autosize_columns()

    def _fill_headings(self):
        row = [column["heading"] for column in self.get_columns()]
        self.append(row)
        self.row_dimensions[1].font = openpyxl.styles.Font(bold=True)

    def _fill_body(self):
        for obj in self.get_queryset():
            self._append_body_row(obj)

    def _append_body_row(self, obj):
        row = []
        for col in self.get_columns():
            if isinstance(col["value"], str):
                row.append(self._get_value_from_dotted_path(obj, col["value"]))
            elif len(inspect.signature(col["value"]).parameters) == 1:
                row.append(col["value"](obj))
            else:
                row.append(col["value"](self, obj))
        self.append(row)

    def _get_value_from_dotted_path(self, obj, attribute_chain):
        result = obj
        for attr in attribute_chain.split("."):
            result = getattr(result, attr)
        return result

    def _autosize_columns(self):
        column_widths = []
        for row in self.iter_rows():
            for i, cell in enumerate(row):
                try:
                    column_widths[i] = max(column_widths[i], len(str(cell.value)))
                except IndexError:
                    column_widths.append(len(str(cell.value)))

        for i, column_width in enumerate(column_widths):
            column = openpyxl.utils.get_column_letter(i + 1)
            self.column_dimensions[column].width = (
                column_width * self.column_width_factor
            )

    @classmethod
    def get_name(self):
        return self.name

    def get_columns(self):
        return self.columns


class WorkbookView(View):
    def get(self, request, *args, **kwargs):
        self._wb = openpyxl.Workbook()
        self._create_sheets()
        self._create_response()
        return self._response

    def _create_sheets(self):
        del self._wb["Sheet"]  # Remove default sheet
        for cls_worksheet in self.worksheets:
            sheet = self._wb.create_sheet(cls_worksheet.get_name())
            sheet.__class__ = cls_worksheet
            sheet.view = self
            sheet.request = self.request
            sheet.fill()

    def _create_response(self):
        obj = BytesIO()
        self._wb.save(obj)
        self._response = HttpResponse(
            obj.getvalue(), content_type=mimetypes.guess_type(self.filename)[0]
        )
        self._response["Content-Disposition"] = f"attachment; filename={self.filename}"
