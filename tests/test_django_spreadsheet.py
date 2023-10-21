from tempfile import NamedTemporaryFile

from django.test import TestCase

import openpyxl

from testapp.models import Author, Book
from testapp.views import AuthorsWorksheet, BooksWorksheet


class WorkbookViewTestCase(TestCase):
    def setUp(self):
        self.response = self.client.get("/downloadbooks/")

    def test_status_code(self):
        self.assertEqual(self.response.status_code, 200)

    def test_filename(self):
        self.assertEqual(
            self.response["Content-Disposition"],
            "attachment; filename=books.xlsx",
        )

    def test_mimetype(self):
        self.assertEqual(
            self.response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_returned_workbook_contains_two_sheets(self):
        with NamedTemporaryFile(suffix=".xlsx") as f:
            f.write(self.response.content)
            wb = openpyxl.load_workbook(f.name)
        self.assertEqual(wb.sheetnames, ["Books", "Authors"])


class GetWorksheetBodyMixin:
    def _get_body(self, worksheet_name):
        worksheet = self.wb[worksheet_name]
        result = []
        first = True
        for row in worksheet.rows:
            if first:
                first = False
                continue
            result.append([cell.value for cell in row])
        return result


class WorksheetTestCase(TestCase, GetWorksheetBodyMixin):
    @classmethod
    def setUpTestData(cls):
        tolkien = Author.objects.create(name="J.R.R.Tolkien")
        rowling = Author.objects.create(name="J.K.Rowling")
        Book.objects.create(author=tolkien, title="The Hobbit")
        Book.objects.create(author=tolkien, title="The Lord of the Rings")
        Book.objects.create(author=rowling, title="The Prisoner of Azkaban")
        Book.objects.create(author=rowling, title="The Goblet of Fire")

    def setUp(self):
        self.response = self.client.get("/downloadbooks/")
        with NamedTemporaryFile(suffix=".xlsx") as f:
            f.write(self.response.content)
            self.wb = openpyxl.load_workbook(f.name)

    def test_headings_in_authors(self):
        first_row = next(self.wb["Authors"].rows)
        self.assertEqual([cell.value for cell in first_row], ["Name", "Name2"])

    def test_headings_in_books(self):
        first_row = next(self.wb["Books"].rows)
        self.assertEqual([cell.value for cell in first_row], ["Author", "Title"])

    def test_body_in_authors(self):
        body = self._get_body("Authors")
        self.assertEqual(
            body,
            [["J.K.Rowling", "J.K.Rowling"], ["J.R.R.Tolkien", "J.R.R.Tolkien"]],
        )

    def test_body_in_books(self):
        body = self._get_body("Books")
        self.assertEqual(
            body,
            [
                ["J.K.Rowling", "The goblet of fire"],
                ["J.K.Rowling", "The prisoner of azkaban"],
                ["J.R.R.Tolkien", "The hobbit"],
                ["J.R.R.Tolkien", "The lord of the rings"],
            ],
        )

    def test_heading_is_bold(self):
        self.assertTrue(self.wb["Books"].row_dimensions[1].font.bold)

    def test_body_is_not_bold(self):
        self.assertFalse(self.wb["Books"].row_dimensions[2].font.bold)

    def test_column_width(self):
        self.assertAlmostEqual(
            self.wb["Books"].column_dimensions["B"].width,
            len("The prisoner of azkaban") * 1.23,
        )


class GetQuerysetTestCase(TestCase, GetWorksheetBodyMixin):
    @classmethod
    def setUpTestData(cls):
        tolkien = Author.objects.create(name="J.R.R.Tolkien")
        rowling = Author.objects.create(name="J.K.Rowling")
        Book.objects.create(author=tolkien, title="The Hobbit")
        Book.objects.create(author=tolkien, title="The Lord of the Rings")
        Book.objects.create(author=rowling, title="The Prisoner of Azkaban")
        Book.objects.create(author=rowling, title="The Goblet of Fire")

    def setUp(self):
        self.response = self.client.get("/downloadbooks/")
        with NamedTemporaryFile(suffix=".xlsx") as f:
            f.write(self.response.content)
            self.wb = openpyxl.load_workbook(f.name)

    def test_body_in_authors(self):
        body = self._get_body("Authors")
        self.assertEqual(
            body,
            [["J.K.Rowling", "J.K.Rowling"], ["J.R.R.Tolkien", "J.R.R.Tolkien"]],
        )


class HasRequestObjectTestCase(TestCase):
    @staticmethod
    def get_queryset(self):
        self.request  # Will raise exception if it does not exist
        return self.model.objects.all()

    def setUp(self):
        self.saved_get_queryset = AuthorsWorksheet.get_queryset
        AuthorsWorksheet.get_queryset = self.get_queryset

    def tearDown(self):
        AuthorsWorksheet.get_queryset = self.saved_get_queryset

    def test_has_request_object(self):
        # The following will raise exception if AuthorsWorksheet has no self.request
        self.response = self.client.get("/downloadbooks/")


class HasViewInstanceTestCase(TestCase):
    @staticmethod
    def get_queryset(self):
        self.view  # Will raise exception if it does not exist
        return self.model.objects.all()

    def setUp(self):
        self.saved_get_queryset = AuthorsWorksheet.get_queryset
        AuthorsWorksheet.get_queryset = self.get_queryset

    def tearDown(self):
        AuthorsWorksheet.get_queryset = self.saved_get_queryset

    def test_has_view_instance(self):
        # The following will raise exception if AuthorsWorksheet has no self.view
        self.response = self.client.get("/downloadbooks/")


class CustomColumnWidthFactorTestCase(TestCase):
    @classmethod
    def setUpTestData(cls):
        tolkien = Author.objects.create(name="J.R.R.Tolkien")
        rowling = Author.objects.create(name="J.K.Rowling")
        Book.objects.create(author=tolkien, title="The Hobbit")
        Book.objects.create(author=tolkien, title="The Lord of the Rings")
        Book.objects.create(author=rowling, title="The Prisoner of Azkaban")
        Book.objects.create(author=rowling, title="The Goblet of Fire")

    def setUp(self):
        self.saved_column_width_factor = BooksWorksheet.column_width_factor
        BooksWorksheet.column_width_factor = 1.42
        self.response = self.client.get("/downloadbooks/")
        with NamedTemporaryFile(suffix=".xlsx") as f:
            f.write(self.response.content)
            self.wb = openpyxl.load_workbook(f.name)

    def tearDown(self):
        BooksWorksheet.column_width_factor = self.saved_column_width_factor

    def test_column_width(self):
        self.assertAlmostEqual(
            self.wb["Books"].column_dimensions["B"].width,
            len("The prisoner of azkaban") * 1.42,
        )
